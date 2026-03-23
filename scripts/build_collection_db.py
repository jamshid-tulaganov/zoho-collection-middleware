#!/usr/bin/env python3

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = ROOT / "collections" / "data" / "TSS_Bad_Debtors_New_25.02.2026.xlsx"
OUTPUT_JSON = ROOT / "collections" / "db" / "collection-placement-db.json"


def to_iso_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text[:10] if len(text) >= 10 else ""


def to_number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def normalize_space(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def normalize_company_key(value):
    return "".join(ch for ch in normalize_space(value).lower() if ch.isalnum())


def build_dataset_rows(workbook):
    ws = workbook["Dataset"]
    grouped = {}

    for row in ws.iter_rows(min_row=9, values_only=True):
        company = normalize_space(row[2])
        debtor_type = normalize_space(row[4])
        placement_date = to_iso_date(row[18])
        company_key = normalize_company_key(company)
        if not company_key:
            continue
        if debtor_type.lower() == "fraud":
            continue

        payload = {
            "invoice_number": str(row[11] or "").strip().replace(".0", ""),
            "company": company,
            "invoice_status": normalize_space(row[3]),
            "debtor_type": debtor_type,
            "collections_agent": normalize_space(row[5]),
            "invoice_date": to_iso_date(row[12]),
            "total_amount": round(to_number(row[13]), 2),
            "total_paid": round(to_number(row[14]), 2),
            "remaining_amount": round(to_number(row[15]), 2),
            "placement_date": placement_date,
            "owner_names": normalize_space(row[23]),
            "phone_numbers": normalize_space(row[24]),
            "emails": normalize_space(row[25]),
            "address": normalize_space(row[26]),
            "state": normalize_space(row[27]),
            "city": normalize_space(row[29]),
            "zip": str(row[30] or "").strip().replace(".0", ""),
            "comments_billing": normalize_space(row[31]),
            "sales_agent": normalize_space(row[32]),
            "source_sheet": "Dataset",
        }
        grouped.setdefault(company_key, []).append(payload)

    return grouped


def main():
    workbook = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    grouped = build_dataset_rows(workbook)
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(grouped, indent=2), encoding="utf-8")
    row_count = sum(len(items) for items in grouped.values())
    print(f"Wrote {row_count} collection rows across {len(grouped)} companies to {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
