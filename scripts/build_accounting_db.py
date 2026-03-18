#!/usr/bin/env python3

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = ROOT / "automation" / "TSS Accounting Spreadsheet(1).xlsx"
OUTPUT_JSON = ROOT / "collections" / "db" / "accounting-client-db.json"

CLIENT_SHEETS = [
    ("Client List (1 billing)", "1_billing"),
    ("Client List 2 Billings", "2_billing"),
    ("Client List (Thursday-Wednesday", "thursday_wednesday"),
]

STATE_CODES = {
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN",
    "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV",
    "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN",
    "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "DC",
}

STREET_SUFFIXES = {
    "ST", "STREET", "RD", "ROAD", "AVE", "AVENUE", "DR", "DRIVE", "LN", "LANE", "BLVD",
    "BOULEVARD", "PL", "PLACE", "CT", "COURT", "CIR", "CIRCLE", "PKWY", "WAY", "TRL",
    "TRAIL", "HWY", "HIGHWAY", "TER", "TERRACE",
}

APT_MARKERS = {"APT", "APT.", "UNIT", "STE", "SUITE", "FLOOR", "FL", "#"}


def normalize_carrier_id(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def normalize_phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) > 10:
        digits = digits[-10:]
    return digits if len(digits) == 10 else ""


def split_name(value):
    parts = [part for part in re.split(r"\s+", str(value or "").strip()) if part]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def normalize_space(value):
    text = str(value or "").replace("\t", " ").replace("\n", " ").strip()
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    text = re.sub(r"([A-Z]{2})(\d{5})(?:-\d{4})?$", r"\1 \2", text)
    text = re.sub(r"(\d)([A-Z][a-z])", r"\1 \2", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_address(raw):
    text = normalize_space(raw)
    if not text:
        return {"raw": "", "line1": "", "city": "", "state": "", "zip": ""}

    match = re.search(r"\b([A-Z]{2})\s+(\d{5})(?:-\d{4})?$", text)
    if not match:
        return {"raw": text, "line1": text, "city": "", "state": "", "zip": ""}

    state = match.group(1)
    zip_code = match.group(2)
    prefix = text[:match.start()].strip()
    tokens = prefix.split()

    city_tokens = []
    index = len(tokens) - 1
    while index >= 0 and len(city_tokens) < 3:
        token = tokens[index]
        upper = re.sub(r"[^A-Z]", "", token.upper())
        if not upper:
            break
        if token.isdigit() or any(ch.isdigit() for ch in token):
            break
        if upper in STATE_CODES or upper in STREET_SUFFIXES or upper in APT_MARKERS:
            break
        city_tokens.insert(0, token)
        index -= 1

    city = " ".join(city_tokens).strip()
    line1 = " ".join(tokens[: index + 1]).strip() if city_tokens else prefix

    if not line1:
        line1 = prefix

    return {
        "raw": text,
        "line1": line1,
        "city": city,
        "state": state,
        "zip": zip_code,
    }


def build_row(sheet_name, billing_cycle, row):
    carrier_id = normalize_carrier_id(row[0])
    if not carrier_id:
        return None

    company = str(row[1] or "").strip()
    full_name = str(row[2] or "").strip()
    email = str(row[3] or "").strip()
    phone_raw = str(row[4] or "").strip()
    credit_score = row[5]
    cards = row[6]
    address = parse_address(row[7])
    agent_name = str(row[8] or "").strip()
    first_name, last_name = split_name(full_name)

    return carrier_id, {
        "carrier_id": carrier_id,
        "source_sheet": sheet_name,
        "billing_cycle_source": billing_cycle,
        "company": company,
        "full_name": full_name,
        "first_name": first_name,
        "last_name": last_name,
        "email": email,
        "phone_raw": phone_raw,
        "phone": normalize_phone(phone_raw),
        "credit_score": int(credit_score) if isinstance(credit_score, (int, float)) else 0,
        "cards": int(cards) if isinstance(cards, (int, float)) else 0,
        "address": address,
        "agent_name": agent_name,
    }


def choose_better(current, candidate):
    current_score = sum(
        bool(current.get(key))
        for key in ("company", "full_name", "phone", "email", "credit_score", "agent_name")
    ) + bool(current.get("address", {}).get("zip"))
    candidate_score = sum(
        bool(candidate.get(key))
        for key in ("company", "full_name", "phone", "email", "credit_score", "agent_name")
    ) + bool(candidate.get("address", {}).get("zip"))
    return candidate if candidate_score > current_score else current


def main():
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    merged = {}

    for sheet_name, billing_cycle in CLIENT_SHEETS:
        ws = workbook[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            built = build_row(sheet_name, billing_cycle, row)
            if not built:
                continue
            carrier_id, payload = built
            if carrier_id in merged:
                merged[carrier_id] = choose_better(merged[carrier_id], payload)
            else:
                merged[carrier_id] = payload

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(merged, indent=2), encoding="utf-8")
    print(f"Wrote {len(merged)} accounting clients to {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
